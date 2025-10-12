import openpyxl
from openpyxl import Workbook
import os
import shutil
from filelock import FileLock
from datetime import datetime
from typing import List, Dict, Any, Optional
import sqlite3

def ensure_excel_file_exists(excel_path: str) -> bool:
    """Ensure Excel file exists with Journal worksheet"""
    try:
        if not os.path.exists(excel_path):
            # Create new workbook with Journal sheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Journal"

            # Add headers
            headers = ["Date", "Category", "Amount", "Memo", "User"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # Create data directory if needed
            os.makedirs(os.path.dirname(excel_path), exist_ok=True)
            wb.save(excel_path)
            return True
        return True
    except Exception:
        return False

def create_backup(excel_path: str) -> bool:
    """Create backup of Excel file"""
    try:
        backup_path = excel_path + ".bak"
        if os.path.exists(excel_path):
            shutil.copy2(excel_path, backup_path)
        return True
    except Exception:
        return False

def get_next_row(worksheet) -> int:
    """Get next available row in worksheet"""
    row = 1
    while worksheet.cell(row=row, column=1).value is not None:
        row += 1
    return row

def append_journal_entry(excel_path: str, category: str, amount: float, memo: str = "", user: str = "") -> bool:
    """Append new entry to Journal worksheet"""
    lock_path = excel_path + ".lock"

    try:
        with FileLock(lock_path):
            # Ensure file exists
            if not ensure_excel_file_exists(excel_path):
                return False

            # Create backup
            create_backup(excel_path)

            # Open workbook
            wb = openpyxl.load_workbook(excel_path)

            # Get or create Journal worksheet
            if "Journal" in wb.sheetnames:
                ws = wb["Journal"]
            else:
                ws = wb.create_sheet("Journal")
                # Add headers if new sheet
                headers = ["Date", "Category", "Amount", "Memo", "User"]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)

            # Find next row
            next_row = get_next_row(ws)

            # Add data
            ws.cell(row=next_row, column=1, value="=TODAY()")  # Date formula
            ws.cell(row=next_row, column=2, value=category)    # Category
            ws.cell(row=next_row, column=3, value=amount)      # Amount
            ws.cell(row=next_row, column=4, value=memo)        # Memo
            ws.cell(row=next_row, column=5, value=user)        # User

            # Save workbook
            wb.save(excel_path)
            wb.close()
            return True

    except Exception as e:
        print(f"Excel append error: {e}")
        return False

def read_journal_balance(excel_path: str) -> Optional[Dict[str, Any]]:
    """Read balance from Excel file (tries to read from Summary sheet if exists)"""
    try:
        if not os.path.exists(excel_path):
            return None

        wb = openpyxl.load_workbook(excel_path, data_only=True)

        # Try to read from Summary sheet first
        if "Summary" in wb.sheetnames:
            summary_ws = wb["Summary"]
            # Look for balance cell (common patterns)
            balance = None

            # Try common balance locations
            for row in range(1, 20):
                for col in range(1, 10):
                    cell_value = summary_ws.cell(row=row, column=col).value
                    if isinstance(cell_value, str) and "balance" in cell_value.lower():
                        # Balance might be in next cell
                        balance_cell = summary_ws.cell(row=row, column=col+1).value
                        if isinstance(balance_cell, (int, float)):
                            balance = balance_cell
                            break
                if balance is not None:
                    break

            if balance is not None:
                wb.close()
                return {
                    "balance": balance,
                    "source": "Summary sheet"
                }

        # Fallback: calculate from Journal
        if "Journal" in wb.sheetnames:
            journal_ws = wb["Journal"]
            total = 0
            count = 0

            # Sum amount column (assuming column 3)
            row = 2  # Skip header
            while journal_ws.cell(row=row, column=3).value is not None:
                amount = journal_ws.cell(row=row, column=3).value
                if isinstance(amount, (int, float)):
                    total += amount
                    count += 1
                row += 1

            wb.close()
            return {
                "balance": total,
                "source": f"Journal calculation ({count} entries)"
            }

        wb.close()
        return None

    except Exception as e:
        print(f"Excel read error: {e}")
        return None

def export_journal_csv(excel_path: str) -> Optional[str]:
    """Export Journal worksheet to CSV string"""
    try:
        if not os.path.exists(excel_path):
            return None

        wb = openpyxl.load_workbook(excel_path, data_only=True)

        if "Journal" not in wb.sheetnames:
            wb.close()
            return None

        ws = wb["Journal"]

        # Convert to CSV
        csv_lines = []
        for row in ws.iter_rows():
            row_data = []
            for cell in row:
                value = cell.value
                if value is None:
                    value = ""
                elif isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d")
                row_data.append(str(value))
            csv_lines.append(",".join(row_data))

        wb.close()
        return "\n".join(csv_lines)

    except Exception as e:
        print(f"Excel export error: {e}")
        return None

def get_guild_excel_path(guild_id: int) -> Optional[str]:
    """Get Excel file path for a guild from configuration"""
    with sqlite3.connect("data/tenant.db") as conn:
        cursor = conn.execute(
            "SELECT excel_path FROM org_configs WHERE guild_id = ?",
            (guild_id,)
        )
        row = cursor.fetchone()
        return row[0] if row and row[0] else None